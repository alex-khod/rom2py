import PIL
from PIL import Image
import openpyxl
from openpyxl import drawing
import io
from src.resources import Resources

"""
import numpy as np
wb = openpyxl.Workbook()
ws = wb.active

arr = np.zeros((20,20), dtype='uint8')
img = Image.fromarray(arr)
    
img = ExportImage(img)
# img = drawing.image.Image(img)

ws.add_image(img, anchor="A1")

wb.save("wb.xlsx")
wb.close()
"""


class ExportImage(drawing.image.Image):
    def _data(self):
        img = self.ref
        fp = io.BytesIO()
        img.save(fp, format="png")
        fp.seek(0)
        data = fp.read()
        fp.close()
        return data


def add_image_column(ws, images, column_id):
    max_width = 0
    image_column_letter = openpyxl.utils.cell.get_column_letter(column_id)

    for i, image in enumerate(images):
        if image is None:
            continue
        max_width = max(max_width, image.width)
        row = i + 2
        image_anchor = image_column_letter + str(row)
        image = ExportImage(image)
        height_in_points = openpyxl.utils.units.pixels_to_points(image.height)
        ws.row_dimensions[row].height = max(18, ws.row_dimensions[row].height or 0, height_in_points)
        ws.add_image(image, anchor=image_anchor)

        ws.column_dimensions[image_column_letter].width = max_width // 6


def transform_coords(x, y):
    MAP_PADDING = 8
    x = x >> 8
    y = y >> 8
    # (unit.x >> 8) + (unit.x & 0xFF) / 256,                   
    return x, y


def coord_check(alm, x, y):
    return (-1 < x < alm.width - 16) and (-1 < y < alm.height - 16)


# assert coord_check(alm, 0,0)
# assert not coord_check(alm, -1,0)
# assert coord_check(alm, alm.width-17,0)
# assert not coord_check(alm, alm.width-16,0)

def coord_check_true(alm, x, y):
    return True


coord_check = coord_check_true


class AlmSheet:

    def __init__(self, alm):
        self.alm = alm


class PlayersSheet(AlmSheet):
    name = 'players'

    def data(self):
        alm = self.alm
        # header = ['unit_id', 'player_id', 'player', 'group_id', 'x', 'y', 'type_id', 'name', 'server_id', 'npcname']
        rel_names = ['foe', 'friend', 'neutral', 'vision']

        players = alm["players"].body.players

        yield ['name'] + [p.name for p in players]
        yield ['color_id'] + [p.color_id for p in players]
        yield ['diplomacy table']

        reltable = [['#'] + [p.name for p in players]]
        for player in players:
            s = [player.name]
            for i, dip in enumerate(player.diplomacy_states[:len(players)]):
                relations = dip & 1, dip & 2, (4 - (dip & 3)) & 4, dip & 16
                relation = ','.join([name for rel, name in zip(relations, rel_names) if rel])
                s.append(relation)
            reltable.append(s)
        for row in reltable:
            yield row


class SacksSheet(AlmSheet):
    name = 'sacks'

    def data(self):
        alm = self.alm
        header = ['money', 'unit_id (host)', 'x', 'y', 'item_count']
        item_header = ['#', 'itemname_no', 'item_id', 'item_name', 'is_magic']
        sacks = alm["sacks"].body.sacks
        itemname_bin = Resources.special('itemname.bin').content
        item_names = Resources["main", "text", "itemname.txt"].content
        yield header
        for sack in sacks:
            x, y = transform_coords(sack.x, sack.y)
            if not coord_check(alm, x, y):
                print(f"Warning: sack id={sack} beyond playzone and was removed")
                continue
            yield [sack.money, sack.unit_id, x, y, sack.item_count]
            if sack.items:
                yield item_header
            for i, item in enumerate(sack.items):
                magic = 'effect_id=%d' % item.effect_id if item.effect_id else 'N/A'
                item_no = itemname_bin[item.item_id]
                item_name = item_names[item_no]
                yield [f'item #{i + 1} ->', item_no, hex(item.item_id), item_name, magic]


class ObjectsSheet(AlmSheet):
    name = 'objects'

    def data(self):
        header = ['x', 'y', 'id', 'desc', 'sprite', 'width', 'height', 'centerx', 'centery']
        yield header

        alm = self.alm
        object_registry = Resources['graphics', 'objects', 'objects.reg'].content
        objects_by_id = object_registry.objects_by_id

        for pos, oid in enumerate(alm["objects"].body.objects):
            if oid > 0:
                x, y = pos % alm.width, pos // alm.width
                try:
                    obj = objects_by_id[oid - 1]
                except KeyError:
                    msg = f"Unknown object_id \"{oid}\" at ({x}, {y})"
                    print("Warning:", msg)
                    # raise IndexError(msg)
                    continue
                x, y = transform_coords(x, y)
                if not coord_check(alm, x, y):
                    continue
                yield [x, y, obj['id'], obj['desctext'], obj['file'], obj['width'], obj['height'], obj['centerx'],
                       obj['centery']]


class StructuresSheet(AlmSheet):
    name = 'structures'

    def data(self):
        header = ['x', 'y', 'id', 'player_id', 'health', 'structure_id(type_id)', 'desc', 'picture(infowindow)',
                  'sprite']
        yield header

        graphics_res = Resources.get('graphics')
        structure_registry = Resources['graphics', 'structures', 'structures.reg'].content
        structures_by_id = structure_registry.structures_by_id
        alm = self.alm
        for obj in alm["structures"].body.structures:
            st = structures_by_id[obj.type_id]
            x, y = obj.x >> 8, obj.y >> 8
            yield [x, y, obj.id, obj.player_id, obj.health, obj.type_id, st['desctext'], st['picture'], st['file']]


class UnitsSheet(AlmSheet):
    name = "units"

    def data(self):
        alm = self.alm
        npc_names = Resources["main", 'text', "npcnames.txt"].content
        unit_names = Resources["main", 'text', "unitname.txt"].content
        header = ['unit_id', 'player_id', 'player_name', 'group_id', 'x', 'y', 'type_id', 'name', 'server_id',
                  'npcname']
        yield header
        players = alm["players"].body.players
        for i, unit in enumerate(alm["units"].body.units):
            if unit.server_id >= 10000:
                print(f"Warning: strange unit name of unit_id={unit.unit_id & 0xFFFF} (srv={unit.server_id})")
                true_name = npc_names[(unit.server_id // 10) & 0xFF]
                print(true_name)
            else:
                true_name = npc_names[unit.server_id - 1]
            x, y = transform_coords(unit.x, unit.y)
            if not coord_check(alm, x, y):
                continue
            row = [unit.unit_id & 0xFFFF,
                   unit.player_id,
                   players[unit.player_id - 1].name,
                   unit.group_id,
                   x,
                   y,
                   unit.type_id, unit_names[unit.type_id],
                   unit.server_id,
                   true_name
                   ]
            yield row


class ChecksSheet(AlmSheet):
    name = 'checks (logic)'

    def data(self):
        header = ["id", "name : type -> args(argName : argType = argValue...)"]
        yield header
        for k, v in enumerate(self.alm["triggers"].body.checks):
            yield [k, str(v)]


class InstancesSheet(AlmSheet):
    name = 'instances (logic)'

    def data(self):
        header = ["id", "name : type -> args(argName : argType = argValue...)"]
        yield header
        for k, v in enumerate(self.alm["triggers"].body.instances):
            yield [k, str(v)]


class TriggersSheet(AlmSheet):
    name = 'triggers (logic)'

    def data(self):
        for trigger in self.alm["triggers"].body.triggers:
            for line in str(trigger).split('\n'):
                yield [line]


def generate_workbook(alm, sheet_cls):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    sheets = [sheet(alm) for sheet in sheet_cls]

    for sheet in sheets:
        ws = wb.create_sheet(sheet.name)
        for row in sheet.data():
            ws.append(row)

        # autofit columns
        for col in ws.columns:
            max_width = max([len(str(cell.value)) for cell in col])
            letter = col[0].column_letter
            ws.column_dimensions[letter].width = 1 + max_width
    return wb
